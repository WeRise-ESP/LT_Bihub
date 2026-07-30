"""
Informe completo Low Ticket — dos meses comparados
5 informes diferenciados por mes:
1. Contactos por País
2. Estados de contacto por País
3. Calidad de Contactos por País (via lifecyclestage)
4. Contactos por Fuente Original de Tráfico
5. Contactos por Fuente Original de Tráfico y País
"""
import requests
import pandas as pd
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from dotenv import load_dotenv
import os

load_dotenv()
console = Console()

TOKEN = os.getenv("HUBSPOT_TOKEN")
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
BASE = "https://api.hubapi.com"

# ── Pipelines de venta Low Ticket ─────────────────────────────────────────────
# La operativa migró de "Pipeline de ventas" (histórico, id `default`) a
# "WooCommerce Orders" en mayo de 2026. El corte evita mezclar los dos mundos
# dentro del mismo período; si el corte real cambia, mueve esta constante.
FECHA_CORTE_WOO = "2026-05-01"

PIPELINE_VENTAS = "default"        # "Pipeline de ventas" (histórico)
PIPELINE_WOO    = "3708462309"     # "WooCommerce Orders" (actual)

LT_STAGES = {
    # Pipeline de ventas (histórico)
    "672731600":  "Lead Ventas",
    "672731601":  "Lead Mkt",
    "672731602":  "Suspendido",
    "672731603":  "En gestión",
    "672731604":  "Hot Lead",
    "672731605":  "Cupón",
    "closedwon":  "Cierre ganado",
    "closedlost": "Cierre perdido",
    "672694235":  "Ajuste post-compra",
    # WooCommerce Orders (actual)
    "5177323744": "Checkout pendiente",
    "5177881838": "Carrito abandonado",
    "5177881839": "En espera",
    "5177881840": "Señal pagada",
    "5177881841": "Pagado / procesando",
    "5146661068": "Completado",
    "5177881842": "Cancelado",
    "5177881843": "Reembolsado",
    "5177881844": "Fallido",
    "5177309408": "Borrador",
}


def pipelines_para(inicio, fin):
    """Pipelines que pueden aportar negocios creados en [inicio, fin]."""
    pipes = []
    if inicio[:10] < FECHA_CORTE_WOO:
        pipes.append(PIPELINE_VENTAS)
    if fin[:10] >= FECHA_CORTE_WOO:
        pipes.append(PIPELINE_WOO)
    return pipes or [PIPELINE_WOO]

# Meses a comparar. Son constantes: edítalas para sacar el informe de otro par
# de meses. Por defecto, los dos últimos del período WooCommerce.
PERIODOS = {
    "Junio": ("2026-06-01T00:00:00Z", "2026-06-30T23:59:59Z"),
    "Julio": ("2026-07-01T00:00:00Z", "2026-07-31T23:59:59Z"),
}

CONTACT_PROPS = [
    "firstname", "lastname", "email",
    "pais_de_residencia", "ip_country", "country",
    "billing_country", "shipping_country",
    "pais_de_la_ip_capabilia", "codigo_de_pais_de_la_ip_capabilia",
    "lt_lead_status", "hs_lead_status", "lifecyclestage",
    "hs_analytics_source", "hs_analytics_source_data_1", "hs_analytics_source_data_2",
    "estado_de_lead_no_valido",
]

FUENTES_ES = {
    "ORGANIC_SEARCH":   "Búsqueda orgánica",
    "PAID_SEARCH":      "Búsqueda pagada",
    "EMAIL_MARKETING":  "Email marketing",
    "SOCIAL_MEDIA":     "Redes sociales",
    "REFERRALS":        "Referencias",
    "OTHER_CAMPAIGNS":  "Otras campañas",
    "DIRECT_TRAFFIC":   "Tráfico directo",
    "OFFLINE":          "Offline",
    "PAID_SOCIAL":      "Social pagado",
    "AI_REFERRALS":     "Referral IA",
    "":                 "Sin datos",
    None:               "Sin datos",
}

LIFECYCLE_ES = {
    "subscriber":             "Suscriptor",
    "lead":                   "Lead",
    "marketingqualifiedlead": "MQL",
    "salesqualifiedlead":     "SQL",
    "opportunity":            "Oportunidad",
    "customer":               "Cliente",
    "evangelist":             "Evangelista",
    "other":                  "Otro",
    "":                       "Sin etapa",
    None:                     "Sin etapa",
}

# Orden de calidad para sorting
LIFECYCLE_ORDER = {
    "Cliente": 1, "Oportunidad": 2, "SQL": 3, "MQL": 4,
    "Lead": 5, "Suscriptor": 6, "Evangelista": 7, "Otro": 8, "Sin etapa": 9,
}


def resolve_pais(cp):
    """Resuelve el país usando múltiples campos en orden de fiabilidad."""
    for campo in [
        "pais_de_residencia",
        "ip_country",
        "pais_de_la_ip_capabilia",
        "country",
        "billing_country",
        "shipping_country",
    ]:
        val = cp.get(campo, "")
        if val and val.strip():
            return val.strip().title()
    return "Sin datos"


# ── Obtener negocios Low Ticket filtrados por mes ─────────────────────────────

def get_lt_deals_periodo(mes, fecha_inicio, fecha_fin):
    """Negocios Low Ticket creados en el período, de los pipelines que aplican."""
    deals = []
    pipes = pipelines_para(fecha_inicio, fecha_fin)
    console.print(f"\n[cyan]Obteniendo negocios Low Ticket — {mes} "
                  f"({fecha_inicio[:10]} → {fecha_fin[:10]}) · "
                  f"pipelines: {', '.join(pipes)}[/cyan]")
    with Progress(SpinnerColumn(), TextColumn(f"  {mes}..."),
                  BarColumn(), TaskProgressColumn(), console=console) as progress:
        task = progress.add_task(mes, total=None)
        for pipeline_id in pipes:
            after = None
            while True:
                payload = {
                    "filterGroups": [{"filters": [
                        {"propertyName": "pipeline",    "operator": "EQ",   "value": pipeline_id},
                        {"propertyName": "createdate",  "operator": "GTE",  "value": fecha_inicio},
                        {"propertyName": "createdate",  "operator": "LTE",  "value": fecha_fin},
                    ]}],
                    "properties": ["dealname", "dealstage", "createdate"],
                    "limit": 100,
                }
                if after:
                    payload["after"] = after
                r = requests.post(f"{BASE}/crm/v3/objects/deals/search",
                                  headers=HEADERS, json=payload)
                data = r.json()
                deals.extend(data.get("results", []))
                progress.update(task, completed=len(deals))
                paging = data.get("paging", {})
                if not paging or not paging.get("next"):
                    break
                after = paging["next"]["after"]
    console.print(f"  [green]{len(deals)} negocios[/green]")
    return deals


# ── Asociaciones deal → contacto ──────────────────────────────────────────────

def get_contact_ids_for_deals(deal_ids):
    deal_to_contact = {}
    with Progress(SpinnerColumn(), TextColumn("  Asociaciones..."),
                  BarColumn(), TaskProgressColumn(), console=console) as progress:
        task = progress.add_task("assoc", total=len(deal_ids))
        for i in range(0, len(deal_ids), 100):
            batch = deal_ids[i:i+100]
            r = requests.post(
                f"{BASE}/crm/v4/associations/deals/contacts/batch/read",
                headers=HEADERS,
                json={"inputs": [{"id": did} for did in batch]},
            )
            if r.status_code == 200:
                for item in r.json().get("results", []):
                    fid = str(item.get("from", {}).get("id", ""))
                    tos = item.get("to", [])
                    if tos:
                        deal_to_contact[fid] = str(tos[0]["toObjectId"])
            progress.update(task, completed=min(i+100, len(deal_ids)))
    return deal_to_contact


# ── Propiedades de contactos en batch ─────────────────────────────────────────

def get_contacts_batch(contact_ids):
    contacts = {}
    unique = list(set(contact_ids))
    with Progress(SpinnerColumn(), TextColumn("  Contactos..."),
                  BarColumn(), TaskProgressColumn(), console=console) as progress:
        task = progress.add_task("contacts", total=len(unique))
        for i in range(0, len(unique), 100):
            batch = unique[i:i+100]
            r = requests.post(f"{BASE}/crm/v3/objects/contacts/batch/read",
                              headers=HEADERS,
                              json={"inputs": [{"id": cid} for cid in batch],
                                    "properties": CONTACT_PROPS})
            if r.status_code == 200:
                for c in r.json().get("results", []):
                    contacts[str(c["id"])] = c["properties"]
            progress.update(task, completed=min(i+100, len(unique)))
    return contacts


# ── DataFrame por mes ─────────────────────────────────────────────────────────

def build_df_mes(deals, deal_to_contact, contacts, mes):
    rows = []
    for deal in deals:
        did = str(deal["id"])
        props = deal["properties"]
        etapa = LT_STAGES.get(props.get("dealstage", ""), props.get("dealstage", ""))
        cid = deal_to_contact.get(did)
        cp = contacts.get(cid, {}) if cid else {}

        pais = resolve_pais(cp)

        fuente_raw = cp.get("hs_analytics_source") or ""
        fuente = FUENTES_ES.get(fuente_raw, fuente_raw or "Sin datos")
        fuente_detalle = cp.get("hs_analytics_source_data_1") or ""

        lifecycle_raw = cp.get("lifecyclestage") or ""
        calidad = LIFECYCLE_ES.get(lifecycle_raw, lifecycle_raw or "Sin etapa")

        rows.append({
            "mes":           mes,
            "deal_id":       did,
            "contact_id":    cid or "",
            "etapa_deal":    etapa,
            "pais":          pais,
            "lead_status":   (cp.get("lt_lead_status")
                              or cp.get("hs_lead_status") or "Sin estado"),
            "calidad":       calidad,
            "fuente":        fuente,
            "fuente_detalle": fuente_detalle,
        })
    return pd.DataFrame(rows)


# ── Estilo Excel ──────────────────────────────────────────────────────────────

def _style_sheet(ws, header_color="1F4E79"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor=header_color)
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="DDDDDD"),
    )
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"


# ── Generar informes en Excel ─────────────────────────────────────────────────

def generar_informes(df):
    os.makedirs("exports", exist_ok=True)
    path = "exports/informe_lt_dos_meses.xlsx"

    meses = ["Junio", "Julio"]
    colores = {"Junio": "1F4E79", "Julio": "2E7D32"}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        for mes in meses:
            dm = df[df["mes"] == mes].copy()
            total_mes = len(dm)
            console.print(f"\n[bold]{mes}[/bold] — {total_mes} deals")

            # ── 1. Contactos por País ──────────────────────────────────────────
            inf1 = (dm.groupby("pais")
                      .size().reset_index(name="Total")
                      .sort_values("Total", ascending=False))
            inf1["% Total"] = (inf1["Total"] / total_mes * 100).round(2)
            inf1.to_excel(writer, sheet_name=f"{mes} - 1 País", index=False)
            _style_sheet(writer.sheets[f"{mes} - 1 País"], colores[mes])

            # ── 2. Estados por País ────────────────────────────────────────────
            inf2 = (dm.groupby(["pais", "lead_status"])
                      .size().reset_index(name="Total")
                      .sort_values(["pais", "Total"], ascending=[True, False]))
            inf2_pivot = (inf2.pivot_table(index="pais", columns="lead_status",
                                           values="Total", fill_value=0, aggfunc="sum"))
            inf2_pivot["TOTAL"] = inf2_pivot.sum(axis=1)
            inf2_pivot = inf2_pivot.sort_values("TOTAL", ascending=False).reset_index()
            inf2_pivot.to_excel(writer, sheet_name=f"{mes} - 2 Estados País", index=False)
            _style_sheet(writer.sheets[f"{mes} - 2 Estados País"], colores[mes])

            # ── 3. Calidad por País ────────────────────────────────────────────
            inf3 = (dm.groupby(["pais", "calidad"])
                      .size().reset_index(name="Total")
                      .sort_values(["pais", "Total"], ascending=[True, False]))
            inf3_pivot = (inf3.pivot_table(index="pais", columns="calidad",
                                           values="Total", fill_value=0, aggfunc="sum"))
            # Ordenar columnas por relevancia
            col_order = [c for c in ["Cliente", "Oportunidad", "SQL", "MQL",
                                      "Lead", "Suscriptor", "Otro", "Sin etapa"]
                         if c in inf3_pivot.columns]
            inf3_pivot = inf3_pivot[col_order]
            inf3_pivot["TOTAL"] = inf3_pivot.sum(axis=1)
            inf3_pivot = inf3_pivot.sort_values("TOTAL", ascending=False).reset_index()
            inf3_pivot.to_excel(writer, sheet_name=f"{mes} - 3 Calidad País", index=False)
            _style_sheet(writer.sheets[f"{mes} - 3 Calidad País"], colores[mes])

            # ── 4. Fuente de tráfico ───────────────────────────────────────────
            inf4 = (dm.groupby(["fuente", "fuente_detalle"])
                      .size().reset_index(name="Total")
                      .sort_values("Total", ascending=False))
            inf4["% Total"] = (inf4["Total"] / total_mes * 100).round(2)
            inf4.to_excel(writer, sheet_name=f"{mes} - 4 Fuente", index=False)
            _style_sheet(writer.sheets[f"{mes} - 4 Fuente"], colores[mes])

            # ── 5. Fuente + País ───────────────────────────────────────────────
            inf5 = (dm.groupby(["pais", "fuente"])
                      .size().reset_index(name="Total")
                      .sort_values(["pais", "Total"], ascending=[True, False]))
            inf5_pivot = (inf5.pivot_table(index="pais", columns="fuente",
                                           values="Total", fill_value=0, aggfunc="sum"))
            inf5_pivot["TOTAL"] = inf5_pivot.sum(axis=1)
            inf5_pivot = inf5_pivot.sort_values("TOTAL", ascending=False).reset_index()
            inf5_pivot.to_excel(writer, sheet_name=f"{mes} - 5 Fuente País", index=False)
            _style_sheet(writer.sheets[f"{mes} - 5 Fuente País"], colores[mes])

        # ── Comparativa Junio vs Julio ──────────────────────────────────────────
        comp = (df.groupby(["mes", "pais"])
                  .size().reset_index(name="Total"))
        comp_pivot = comp.pivot_table(index="pais", columns="mes",
                                       values="Total", fill_value=0, aggfunc="sum")
        for m in meses:
            if m not in comp_pivot.columns:
                comp_pivot[m] = 0
        comp_pivot = comp_pivot[meses]
        comp_pivot["Variación"] = comp_pivot.get("Julio", 0) - comp_pivot.get("Junio", 0)
        comp_pivot["Var %"] = ((comp_pivot["Variación"] /
                                comp_pivot["Junio"].replace(0, 1)) * 100).round(1)
        comp_pivot["TOTAL"] = comp_pivot["Junio"] + comp_pivot["Julio"]
        comp_pivot = comp_pivot.sort_values("TOTAL", ascending=False).reset_index()
        comp_pivot.to_excel(writer, sheet_name="COMPARATIVA Jun vs Jul", index=False)
        _style_sheet(writer.sheets["COMPARATIVA Jun vs Jul"], "4A148C")

        # ── Datos completos ────────────────────────────────────────────────────
        df.to_excel(writer, sheet_name="DATOS COMPLETOS", index=False)
        _style_sheet(writer.sheets["DATOS COMPLETOS"], "37474F")

    console.print(f"\n[bold green]Informe generado: {path}[/bold green]")
    return path


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    console.print("\n[bold blue]══════════════════════════════════════════[/bold blue]")
    console.print("[bold blue]  INFORME LOW TICKET — BIHUB  [/bold blue]")
    console.print("[bold blue]══════════════════════════════════════════[/bold blue]")

    all_dfs = []

    for mes, (inicio, fin) in PERIODOS.items():
        deals = get_lt_deals_periodo(mes, inicio, fin)
        if not deals:
            console.print(f"[yellow]Sin negocios en {mes}[/yellow]")
            continue

        deal_ids = [d["id"] for d in deals]
        deal_to_contact = get_contact_ids_for_deals(deal_ids)
        contact_ids = list(set(deal_to_contact.values()))
        contacts = get_contacts_batch(contact_ids)

        dm = build_df_mes(deals, deal_to_contact, contacts, mes)
        all_dfs.append(dm)

        # Preview rápido
        console.print(f"\n  Top países {mes}: " +
                      ", ".join(f"{r['pais']} ({r['Total']})"
                                for _, r in dm.groupby("pais").size()
                                              .reset_index(name="Total")
                                              .sort_values("Total", ascending=False)
                                              .head(5).iterrows()))

    if not all_dfs:
        console.print("[red]No hay datos en ningún período.[/red]")
        return None

    df = pd.concat(all_dfs, ignore_index=True)
    console.print(f"\n[green]Total registros combinados: {len(df)} "
                  f"(Junio: {len(df[df['mes']=='Junio'])} | "
                  f"Julio: {len(df[df['mes']=='Julio'])})[/green]")

    return generar_informes(df)


if __name__ == "__main__":
    run()
