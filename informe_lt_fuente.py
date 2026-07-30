"""
Informe: Estado de Lead por Fuente de Tráfico — Low Ticket
Junio y Julio diferenciados.

Fuente: hs_analytics_source (original). Si vacío → hs_latest_source (más reciente).
Misma lógica de estados que informe_lt_lead_status.py
"""
import requests
import pandas as pd
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()
console = Console()

TOKEN = os.getenv("HUBSPOT_TOKEN")
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
BASE = "https://api.hubapi.com"

# ── Pipelines de venta Low Ticket ─────────────────────────────────────────────
# La operativa migró de "Pipeline de ventas" (histórico, id `default`) a
# "WooCommerce Orders" en mayo de 2026. El corte evita mezclar los dos mundos
# dentro del mismo período; si el corte real cambia, mueve esta constante.
FECHA_CORTE_WOO = "2026-05-01"

PIPELINE_VENTAS = "default"        # "Pipeline de ventas" (histórico)
PIPELINE_WOO    = "3708462309"     # "WooCommerce Orders" (actual)

LT_STAGES = {
    # Pipeline de ventas (histórico)
    "672731600":  "Lead Ventas",
    "672731601":  "Lead Mkt",
    "672731602":  "Suspendido",
    "672731603":  "En gestión",
    "672731604":  "Hot Lead",
    "672731605":  "Cupón",
    "closedwon":  "Cierre ganado",
    "closedlost": "Cierre perdido",
    "672694235":  "Ajuste post-compra",
    # WooCommerce Orders (actual)
    "5177323744": "Checkout pendiente",
    "5177881838": "Carrito abandonado",
    "5177881839": "En espera",
    "5177881840": "Señal pagada",
    "5177881841": "Pagado / procesando",
    "5146661068": "Completado",
    "5177881842": "Cancelado",
    "5177881843": "Reembolsado",
    "5177881844": "Fallido",
    "5177309408": "Borrador",
}


def pipelines_para(inicio, fin):
    """Pipelines que pueden aportar negocios creados en [inicio, fin]."""
    pipes = []
    if inicio[:10] < FECHA_CORTE_WOO:
        pipes.append(PIPELINE_VENTAS)
    if fin[:10] >= FECHA_CORTE_WOO:
        pipes.append(PIPELINE_WOO)
    return pipes or [PIPELINE_WOO]

# Meses a comparar. Son constantes: edítalas para sacar el informe de otro par
# de meses. Por defecto, los dos últimos del período WooCommerce.
PERIODOS = {
    "Junio": ("2026-06-01T00:00:00Z", "2026-06-30T23:59:59Z"),
    "Julio": ("2026-07-01T00:00:00Z", "2026-07-31T23:59:59Z"),
}

CONTACT_PROPS = [
    "firstname", "lastname", "email",
    "lt_lead_status", "hs_lead_status",
    "num_contacted_notes",
    "estado_de_lead_no_valido",
    "hs_analytics_source",
    "hs_analytics_source_data_1",
    "hs_latest_source",
    "hs_latest_source_data_1",
]

DEAL_PROPS = [
    "dealname", "dealstage", "createdate", "amount", "curso",
]
# Etapas de los pipelines que representan una venta perdida. Low Ticket no tiene
# una propiedad de motivo de cierre perdido, así que el motivo ES la etapa en la
# que murió el pedido.
ETAPAS_PERDIDA = {
    "Cierre perdido", "Cancelado", "Fallido", "Reembolsado",
    "Carrito abandonado", "Suspendido",
}


FUENTES_ES = {
    "ORGANIC_SEARCH":  "Búsqueda orgánica",
    "PAID_SEARCH":     "Búsqueda pagada",
    "EMAIL_MARKETING": "Email marketing",
    "SOCIAL_MEDIA":    "Redes sociales",
    "REFERRALS":       "Referencias",
    "OTHER_CAMPAIGNS": "Otras campañas",
    "DIRECT_TRAFFIC":  "Tráfico directo",
    "OFFLINE":         "Offline",
    "PAID_SOCIAL":     "Social pagado",
    "AI_REFERRALS":    "Referral IA",
}

LEAD_STATUS_NORM = {
    "nuevo":                          "Nuevo",
    "new":                            "Nuevo",
    "primera respuesta automatizada": "Primera respuesta automatizada",
    "conversación iniciada":          "Conversación iniciada",
    "conversacion iniciada":          "Conversación iniciada",
    "negocio abierto":                "Negocio abierto",
    "negocio ganado":                 "Negocio ganado",
}

ESTADOS_ORDEN = [
    "Nuevo", "Primera respuesta automatizada", "Conversación iniciada",
    "Negocio abierto", "Negocio ganado", "Sin estado",
]

COLORES_ESTADO = {
    "Nuevo":                          "C8E6C9",
    "Primera respuesta automatizada": "FFF9C4",
    "Conversación iniciada":          "B3E5FC",
    "Negocio abierto":                "FFE0B2",
    "Negocio ganado":                 "80CBC4",
    "Sin estado":                     "EEEEEE",
}

COLORES_FUENTE = {
    "Búsqueda orgánica": "E8F5E9",
    "Búsqueda pagada":   "E3F2FD",
    "Email marketing":   "FFF3E0",
    "Redes sociales":    "FCE4EC",
    "Social pagado":     "F3E5F5",
    "Tráfico directo":   "E0F7FA",
    "Otras campañas":    "FFF8E1",
    "Offline":           "EFEBE9",
    "Referencias":       "F1F8E9",
    "Referral IA":       "EDE7F6",
    "Sin datos":         "FAFAFA",
}


def resolve_fuente(cp):
    """Original si tiene valor, si no → más reciente. Siempre con etiqueta de origen."""
    raw_orig = (cp.get("hs_analytics_source") or "").strip()
    raw_rec  = (cp.get("hs_latest_source") or "").strip()

    if raw_orig:
        fuente = FUENTES_ES.get(raw_orig, raw_orig.replace("_", " ").title())
        origen = "Original"
        detalle = (cp.get("hs_analytics_source_data_1") or "").strip()
    elif raw_rec:
        fuente = FUENTES_ES.get(raw_rec, raw_rec.replace("_", " ").title())
        origen = "Más reciente"
        detalle = (cp.get("hs_latest_source_data_1") or "").strip()
    else:
        fuente = "Sin datos"
        origen = "—"
        detalle = ""

    return fuente, origen, detalle


# Estados del embudo High Ticket (hs_lead_status) traducidos al embudo Low
# Ticket. Se usan solo como fallback: una parte de los contactos LT los trabaja
# el equipo comercial sobre el embudo antiguo y no tienen lt_lead_status.
HS_A_LT = {
    "matriculado":          "Negocio ganado",
    "negocio cerrado":      "Negocio abierto",
    "contactado":           "Conversación iniciada",
    "intentando contactar": "Primera respuesta automatizada",
}


def norm_lead_status(raw):
    """Etiqueta canónica del embudo Low Ticket. Todo lo que no encaje en el
    embudo (estados de descarte de High Ticket, valores sueltos) cae en
    "Sin estado" para no ensuciar las tablas con etiquetas de otro embudo."""
    if not raw:
        return "Sin estado"
    k = raw.lower().strip()
    if k in LEAD_STATUS_NORM:
        return LEAD_STATUS_NORM[k]
    return HS_A_LT.get(k, "Sin estado")


# ── Fetch ─────────────────────────────────────────────────────────────────────

def get_lt_deals(mes, inicio, fin):
    """Negocios Low Ticket creados en el período, de los pipelines que aplican."""
    deals = []
    for pipeline_id in pipelines_para(inicio, fin):
        after = None
        with Progress(SpinnerColumn(), TextColumn(f"  Negocios {mes}..."),
                      BarColumn(), TaskProgressColumn(), console=console) as p:
            task = p.add_task(mes, total=None)
            while True:
                payload = {
                    "filterGroups": [{"filters": [
                        {"propertyName": "pipeline",   "operator": "EQ",  "value": pipeline_id},
                        {"propertyName": "createdate", "operator": "GTE", "value": inicio},
                        {"propertyName": "createdate", "operator": "LTE", "value": fin},
                    ]}],
                    "properties": DEAL_PROPS,
                    "limit": 100,
                }
                if after:
                    payload["after"] = after
                r = requests.post(f"{BASE}/crm/v3/objects/deals/search",
                                  headers=HEADERS, json=payload)
                data = r.json()
                deals.extend(data.get("results", []))
                p.update(task, completed=len(deals))
                paging = data.get("paging", {})
                if not paging or not paging.get("next"):
                    break
                after = paging["next"]["after"]
    return deals


def get_contact_ids(deal_ids):
    deal_to_contact = {}
    with Progress(SpinnerColumn(), TextColumn("  Asociaciones..."),
                  BarColumn(), TaskProgressColumn(), console=console) as p:
        task = p.add_task("assoc", total=len(deal_ids))
        for i in range(0, len(deal_ids), 100):
            batch = deal_ids[i:i+100]
            r = requests.post(
                f"{BASE}/crm/v4/associations/deals/contacts/batch/read",
                headers=HEADERS,
                json={"inputs": [{"id": d} for d in batch]},
            )
            if r.status_code == 200:
                for item in r.json().get("results", []):
                    fid = str(item.get("from", {}).get("id", ""))
                    tos = item.get("to", [])
                    if tos:
                        deal_to_contact[fid] = str(tos[0]["toObjectId"])
            p.update(task, completed=min(i+100, len(deal_ids)))
    return deal_to_contact


def get_contacts(contact_ids):
    contacts = {}
    unique = list(set(contact_ids))
    with Progress(SpinnerColumn(), TextColumn("  Contactos..."),
                  BarColumn(), TaskProgressColumn(), console=console) as p:
        task = p.add_task("contacts", total=len(unique))
        for i in range(0, len(unique), 100):
            batch = unique[i:i+100]
            r = requests.post(f"{BASE}/crm/v3/objects/contacts/batch/read",
                              headers=HEADERS,
                              json={"inputs": [{"id": c} for c in batch],
                                    "properties": CONTACT_PROPS})
            if r.status_code == 200:
                for c in r.json().get("results", []):
                    contacts[str(c["id"])] = c["properties"]
            p.update(task, completed=min(i+100, len(unique)))
    return contacts


def build_df(deals, deal_to_contact, contacts, mes):
    rows = []
    for deal in deals:
        did = str(deal["id"])
        dp = deal["properties"]
        etapa = LT_STAGES.get(dp.get("dealstage", ""), dp.get("dealstage", ""))
        motivo_cierre = etapa if etapa in ETAPAS_PERDIDA else ""

        cid = deal_to_contact.get(did)
        cp = contacts.get(cid, {}) if cid else {}

        fuente, origen_fuente, fuente_detalle = resolve_fuente(cp)
        lead_status = norm_lead_status(cp.get("lt_lead_status")
                                   or cp.get("hs_lead_status") or "")
        intentos = int(cp.get("num_contacted_notes") or 0)
        motivo_no_valido = cp.get("estado_de_lead_no_valido") or ""

        rows.append({
            "mes":              mes,
            "deal_id":          did,
            "fuente":           fuente,
            "origen_fuente":    origen_fuente,
            "fuente_detalle":   fuente_detalle,
            "lead_status":      lead_status,
            "intentos":         intentos,
            "motivo_no_valido": motivo_no_valido,
            "motivo_cierre":    motivo_cierre,
            "etapa_deal":       etapa,
        })
    return pd.DataFrame(rows)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def style_header(ws, color="1F4E79"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 38)
    ws.freeze_panes = "B2"


def color_fuente_rows(ws, fuente_col_idx):
    for row in ws.iter_rows(min_row=2):
        fuente = str(row[fuente_col_idx].value or "")
        hex_col = COLORES_FUENTE.get(fuente, "FFFFFF")
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=hex_col)
            cell.alignment = Alignment(horizontal="center")


# ── Hojas de informe ──────────────────────────────────────────────────────────

def hoja_resumen(df, writer, mes, color):
    """Pivot principal: Fuente × Estado de Lead."""
    dm = df[df["mes"] == mes]
    total = len(dm)

    pivot = (dm.groupby(["fuente", "lead_status"])
               .size()
               .reset_index(name="Total"))
    pivot_wide = pivot.pivot_table(index="fuente", columns="lead_status",
                                   values="Total", fill_value=0, aggfunc="sum")
    for estado in ESTADOS_ORDEN:
        if estado not in pivot_wide.columns:
            pivot_wide[estado] = 0
    pivot_wide = pivot_wide[[e for e in ESTADOS_ORDEN if e in pivot_wide.columns]]
    pivot_wide["TOTAL"] = pivot_wide.sum(axis=1)
    pivot_wide["% del Total"] = (pivot_wide["TOTAL"] / total * 100).round(1)
    pivot_wide = pivot_wide.sort_values("TOTAL", ascending=False).reset_index()
    pivot_wide.rename(columns={"fuente": "Fuente de Tráfico"}, inplace=True)

    sname = f"{mes} - Resumen por Fuente"
    pivot_wide.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)

    # Colorear columnas de estados
    header_map = {cell.value: cell.column for cell in ws[1]}
    for estado, hex_col in COLORES_ESTADO.items():
        if estado in header_map:
            col_idx = header_map[estado]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value and cell.value > 0:
                        cell.fill = PatternFill("solid", fgColor=hex_col)
                    cell.alignment = Alignment(horizontal="center")

    # Colorear columna fuente
    f_col = header_map.get("Fuente de Tráfico")
    if f_col:
        for row in ws.iter_rows(min_row=2, min_col=f_col, max_col=f_col):
            for cell in row:
                fuente = str(cell.value or "")
                cell.fill = PatternFill("solid", fgColor=COLORES_FUENTE.get(fuente, "FFFFFF"))
                cell.font = Font(bold=True)


def hoja_origen_fuente(df, writer, mes, color):
    """Desglose de cuántos vienen de fuente original vs más reciente."""
    dm = df[df["mes"] == mes]
    resumen = (dm.groupby(["fuente", "origen_fuente"])
                 .size().reset_index(name="Total"))
    pivot = resumen.pivot_table(index="fuente", columns="origen_fuente",
                                values="Total", fill_value=0, aggfunc="sum")
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False).reset_index()
    pivot.rename(columns={"fuente": "Fuente"}, inplace=True)

    sname = f"{mes} - Origen Fuente"
    pivot.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)

    header_map = {cell.value: cell.column for cell in ws[1]}
    f_col = header_map.get("Fuente")
    for row in ws.iter_rows(min_row=2):
        fuente = str(row[0].value or "")
        bg = COLORES_FUENTE.get(fuente, "FFFFFF")
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")


def hoja_intentos(df, writer, mes, color):
    """Intentando contactar: fuente, nº contactos, intentos totales y promedio."""
    dm = df[(df["mes"] == mes) & (df["lead_status"] == "Intentando contactar")]
    if dm.empty:
        return
    resumen = (dm.groupby("fuente")
                 .agg(Contactos=("deal_id","count"),
                      Total_intentos=("intentos","sum"),
                      Promedio_intentos=("intentos","mean"),
                      Max_intentos=("intentos","max"))
                 .round(1).sort_values("Contactos", ascending=False).reset_index())
    resumen.columns = ["Fuente", "Nº Contactos", "Total Intentos",
                        "Promedio Intentos", "Máx Intentos"]
    sname = f"{mes} - Intentando Contactar"[:31]
    resumen.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)
    fill = PatternFill("solid", fgColor=COLORES_ESTADO["Intentando contactar"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def hoja_no_valido(df, writer, mes, color):
    """No válido: fuente × motivo de invalidación."""
    dm = df[(df["mes"] == mes) & (df["lead_status"] == "No válido")]
    if dm.empty:
        return
    dm = dm.copy()
    dm["motivo_no_valido"] = dm["motivo_no_valido"].replace("", "Sin motivo especificado")
    pivot = (dm.groupby(["fuente", "motivo_no_valido"])
               .size().reset_index(name="Total"))
    pivot_wide = pivot.pivot_table(index="fuente", columns="motivo_no_valido",
                                   values="Total", fill_value=0, aggfunc="sum")
    pivot_wide["TOTAL"] = pivot_wide.sum(axis=1)
    pivot_wide = pivot_wide.sort_values("TOTAL", ascending=False).reset_index()
    pivot_wide.rename(columns={"fuente": "Fuente"}, inplace=True)

    sname = f"{mes} - No Válido (Motivos)"[:31]
    pivot_wide.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)
    fill = PatternFill("solid", fgColor=COLORES_ESTADO["No válido"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def hoja_cierre_perdido(df, writer, mes, color):
    """Negocios perdidos: fuente × etapa en la que se perdió el pedido."""
    dm = df[(df["mes"] == mes) & (df["etapa_deal"].isin(ETAPAS_PERDIDA))]
    if dm.empty:
        return
    dm = dm.copy()
    dm["motivo_cierre"] = dm["motivo_cierre"].replace("", "Sin etapa registrada")
    pivot = (dm.groupby(["fuente", "motivo_cierre"])
               .size().reset_index(name="Total"))
    pivot_wide = pivot.pivot_table(index="fuente", columns="motivo_cierre",
                                   values="Total", fill_value=0, aggfunc="sum")
    pivot_wide["TOTAL"] = pivot_wide.sum(axis=1)
    pivot_wide = pivot_wide.sort_values("TOTAL", ascending=False).reset_index()
    pivot_wide.rename(columns={"fuente": "Fuente"}, inplace=True)

    sname = f"{mes} - Negocios perdidos"[:31]
    pivot_wide.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)
    fill = PatternFill("solid", fgColor="FFCDD2")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def hoja_detalle(df, writer, mes, color):
    """Detalle completo con fuente, origen y estado."""
    dm = df[df["mes"] == mes][[
        "fuente", "origen_fuente", "fuente_detalle", "lead_status",
        "intentos", "motivo_no_valido", "motivo_cierre", "etapa_deal",
    ]].sort_values(["fuente", "lead_status"])

    sname = f"{mes} - Detalle Completo"
    dm.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, color)
    autofit(ws)

    header_map = {cell.value: cell.column for cell in ws[1]}
    status_col = header_map.get("lead_status")
    for row in ws.iter_rows(min_row=2):
        estado = str(row[status_col - 1].value or "") if status_col else ""
        hex_col = COLORES_ESTADO.get(estado, "FFFFFF")
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=hex_col)
            cell.alignment = Alignment(horizontal="left")


def hoja_comparativa(df, writer):
    """Comparativa Junio vs Julio por fuente y estado."""
    comp = (df.groupby(["mes", "fuente", "lead_status"])
              .size().reset_index(name="Total"))
    pivot = comp.pivot_table(index=["fuente", "lead_status"], columns="mes",
                              values="Total", fill_value=0, aggfunc="sum").reset_index()
    for m in ["Junio", "Julio"]:
        if m not in pivot.columns:
            pivot[m] = 0
    pivot["Variación"] = pivot["Julio"] - pivot["Junio"]
    pivot["Var %"] = ((pivot["Variación"] / pivot["Junio"].replace(0, 1)) * 100).round(1)
    pivot.rename(columns={"fuente": "Fuente", "lead_status": "Estado Lead"}, inplace=True)
    pivot = pivot.sort_values(["Fuente", "Estado Lead"])

    sname = "COMPARATIVA Jun vs Jul"
    pivot.to_excel(writer, sheet_name=sname, index=False)
    ws = writer.sheets[sname]
    style_header(ws, "4A148C")
    autofit(ws)

    header_map = {cell.value: cell.column for cell in ws[1]}
    var_col = header_map.get("Variación")
    if var_col:
        for row in ws.iter_rows(min_row=2):
            val = row[var_col - 1].value
            if val and val > 0:
                row[var_col - 1].fill = PatternFill("solid", fgColor="C8E6C9")
            elif val and val < 0:
                row[var_col - 1].fill = PatternFill("solid", fgColor="FFCDD2")


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    console.print("\n[bold blue]══════════════════════════════════════════════════[/bold blue]")
    console.print("[bold blue]  INFORME FUENTE DE TRÁFICO × ESTADO — LOW TICKET [/bold blue]")
    console.print("[bold blue]══════════════════════════════════════════════════[/bold blue]\n")

    colores = {"Junio": "1565C0", "Julio": "2E7D32"}
    all_dfs = []

    for mes, (inicio, fin) in PERIODOS.items():
        console.print(f"\n[bold cyan]── {mes.upper()} ──[/bold cyan]")
        deals = get_lt_deals(mes, inicio, fin)
        if not deals:
            console.print(f"[yellow]Sin deals en {mes}[/yellow]")
            continue

        deal_to_contact = get_contact_ids([d["id"] for d in deals])
        contacts = get_contacts(list(set(deal_to_contact.values())))
        dm = build_df(deals, deal_to_contact, contacts, mes)
        all_dfs.append(dm)

        # Preview
        dist_fuente = dm["fuente"].value_counts()
        console.print(f"  Fuentes {mes}: " +
                      " | ".join(f"{k}: {v}" for k, v in dist_fuente.items()))
        n_reciente = (dm["origen_fuente"] == "Más reciente").sum()
        n_original = (dm["origen_fuente"] == "Original").sum()
        n_sin = (dm["origen_fuente"] == "—").sum()
        console.print(f"  Origen fuente → Original: {n_original} | "
                      f"Más reciente (fallback): {n_reciente} | Sin datos: {n_sin}")

    if not all_dfs:
        console.print("[red]Sin datos.[/red]")
        return

    df = pd.concat(all_dfs, ignore_index=True)
    os.makedirs("exports", exist_ok=True)
    path = "exports/informe_lt_fuente_trafico.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for mes in ["Junio", "Julio"]:
            if mes not in df["mes"].values:
                continue
            color = colores[mes]
            hoja_resumen(df, writer, mes, color)
            hoja_origen_fuente(df, writer, mes, color)
            hoja_intentos(df, writer, mes, color)
            hoja_no_valido(df, writer, mes, color)
            hoja_cierre_perdido(df, writer, mes, color)
            hoja_detalle(df, writer, mes, color)

        hoja_comparativa(df, writer)

        df.to_excel(writer, sheet_name="DATOS COMPLETOS", index=False)
        style_header(writer.sheets["DATOS COMPLETOS"], "37474F")
        autofit(writer.sheets["DATOS COMPLETOS"])

    console.print(f"\n[bold green]Informe generado: {path}[/bold green]")
    console.print(f"[dim]Total: {len(df)} registros "
                  f"(Junio: {len(df[df['mes']=='Junio'])} | "
                  f"Julio: {len(df[df['mes']=='Julio'])})[/dim]")
    return path


if __name__ == "__main__":
    run()
